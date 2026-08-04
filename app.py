import email
import os
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    flash,
    Response,
    send_file,
    make_response
)
import mysql.connector
from mysql.connector import Error
#from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import csv
import io

app = Flask(__name__)

UPLOAD_FOLDER = "static/uploads/books"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

def allowed_file(filename):

    return (

        "." in filename

        and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

    )

app.secret_key = "super_secret_library_key"


# ------------------------------
# DATABASE CONNECTION
# ------------------------------
def get_db_connection():
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="prince21",
            database="library_db"
        )
        return conn
    except Error as e:
        print("Database Connection Error:", e)
        return None

def log_activity(message):

    conn = get_db_connection()

    if conn is None:
        return

    cursor = conn.cursor()

    cursor.execute(

        "INSERT INTO activity_logs(activity) VALUES(%s)",

        (message,)

    )

    conn.commit()

    cursor.close()

    conn.close()    


# ------------------------------
# HOME PAGE
# ------------------------------
@app.route('/')
def login_page():
    return render_template("login.html")


# ------------------------------
# LOGIN
# ------------------------------
@app.route('/login', methods=['POST'])
def login():

    role = request.form.get("role")
    username = request.form.get("username")
    password = request.form.get("password")

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("login_page"))

    cursor = conn.cursor(dictionary=True)

    try:

        if role == "admin":

            cursor.execute(
                "SELECT * FROM admins WHERE username=%s",
                (username,)
            )

            admin = cursor.fetchone()
            #print("Admin:", admin)

            if admin and admin["password"] == password:

                session["user_id"] = admin["id"]
                session["role"] = "admin"
                session["name"] = admin["username"]
                session["theme"] = admin["theme"]

                return redirect(url_for("admin_dashboard"))

        elif role == "student":

            cursor.execute(
                "SELECT * FROM students WHERE id=%s",
                (username,)
            )

            student = cursor.fetchone()
            #print("Student:", student)

            if student and student["password"] == password:

                session["user_id"] = student["id"]
                session["role"] = "student"
                session["name"] = student["name"]
                session["theme"] = student["theme"]

                return redirect(url_for("student_dashboard"))

        flash("Invalid username or password.")
        return redirect(url_for("login_page"))

    finally:
        cursor.close()
        conn.close()


# ------------------------------
# LOGOUT
# ------------------------------
@app.route("/logout")
def logout():

    session.clear()
    #flash("Logged out successfully.")
    return redirect(url_for("login_page"))

# ------------------------------
# SAVE THEME
# ------------------------------
@app.route("/save-theme", methods=["POST"])
def save_theme():

    if "role" not in session:
        return {"success": False}, 401

    theme = request.form.get("theme")

    if theme not in ["light", "dark"]:
        return {"success": False}, 400

    conn = get_db_connection()

    if conn is None:
        return {"success": False}, 500

    cursor = conn.cursor()

    try:

        if session["role"] == "admin":

            cursor.execute(
                """
                UPDATE admins
                SET theme=%s
                WHERE id=%s
                """,
                (
                    theme,
                    session["user_id"]
                )
            )

        else:

            cursor.execute(
                """
                UPDATE students
                SET theme=%s
                WHERE id=%s
                """,
                (
                    theme,
                    session["user_id"]
                )
            )

        conn.commit()

        session["theme"] = theme

        return {"success": True}

    finally:

        cursor.close()
        conn.close()


# ------------------------------
# ADMIN DASHBOARD
# ------------------------------
@app.route("/admin/dashboard")
def admin_dashboard():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("login_page"))

    cursor = conn.cursor(dictionary=True)

    try:

        # Total Books
        cursor.execute("SELECT COUNT(*) AS total FROM books")
        total_books = cursor.fetchone()["total"]

        # Total Students
        cursor.execute("SELECT COUNT(*) AS total FROM students")
        total_students = cursor.fetchone()["total"]

        # Issued Books
        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM issued_books
            WHERE status='Issued'
        """)
        issued_books = cursor.fetchone()["total"]

        # Overdue Books
        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM issued_books
            WHERE status='Issued'
            AND return_date < CURDATE()
        """)
        overdue_books = cursor.fetchone()["total"]
        
        cursor.execute("""
        
        SELECT *
        
        FROM activity_logs
        
        ORDER BY id DESC
        
        LIMIT 5
        
        """)
        
        activities = cursor.fetchall()

        return render_template(
            "admin_dashboard.html",
            name=session["name"],
            total_books=total_books,
            total_students=total_students,
            issued_books=issued_books,
            overdue_books=overdue_books,
            activities=activities
        )

    finally:

        cursor.close()
        conn.close()

# ------------------------------
# ADMIN PROFILE
# ------------------------------
@app.route("/admin/profile", methods=["GET", "POST"])
def admin_profile():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("admin_dashboard"))

    cursor = conn.cursor(dictionary=True)

    try:

        if request.method == "POST":

            name = request.form.get("name").strip()
            username = request.form.get("username").strip()
            email = request.form.get("email").strip()

            image = request.files.get("profile_image")

            current_password = request.form.get("current_password", "").strip()
            new_password = request.form.get("new_password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()

            if not name or not username or not email:

                flash("All fields are required.")

                return redirect(url_for("admin_profile"))

            # ------------------------------
            # Upload Profile Picture
            # ------------------------------

            filename = None

            if image and image.filename != "":

                extension = image.filename.rsplit(".", 1)[1].lower()

                filename = f"admin.{extension}"

                image.save(
                    os.path.join(
                        app.static_folder,
                        "uploads",
                        "profiles",
                        filename
                    )
                )

            # ------------------------------
            # Change Password
            # ------------------------------

            password = None

            if current_password or new_password or confirm_password:

                if not current_password or not new_password or not confirm_password:

                    flash("Please fill all password fields.")

                    return redirect(url_for("admin_profile"))

                cursor.execute("""
                    SELECT password
                    FROM admins
                    WHERE id=%s
                """,
                (
                    1,
                ))

                admin_data = cursor.fetchone()

                if admin_data["password"] != current_password:

                    flash("Current password is incorrect.")

                    return redirect(url_for("admin_profile"))

                if new_password != confirm_password:

                    flash("New password and confirm password do not match.")

                    return redirect(url_for("admin_profile"))

                password = new_password

            # ------------------------------
            # Update Profile
            # ------------------------------

            if filename and password:

                cursor.execute("""
                    UPDATE admins
                    SET
                        name=%s,
                        username=%s,
                        email=%s,
                        password=%s,
                        profile_image=%s
                    WHERE id=%s
                """,
                (
                    name,
                    username,
                    email,
                    password,
                    filename,
                    1
                ))

            elif filename:

                cursor.execute("""
                    UPDATE admins
                    SET
                        name=%s,
                        username=%s,
                        email=%s,
                        profile_image=%s
                    WHERE id=%s
                """,
                (
                    name,
                    username,
                    email,
                    filename,
                    1
                ))

            elif password:

                cursor.execute("""
                    UPDATE admins
                    SET
                        name=%s,
                        username=%s,
                        email=%s,
                        password=%s
                    WHERE id=%s
                """,
                (
                    name,
                    username,
                    email,
                    password,
                    1
                ))

            else:

                cursor.execute("""
                    UPDATE admins
                    SET
                        name=%s,
                        username=%s,
                        email=%s
                    WHERE id=%s
                """,
                (
                    name,
                    username,
                    email,
                    1
                ))

            conn.commit()

            session["name"] = name
            session["username"] = username

            if password:

                flash("Profile and password updated successfully.")

            else:

                flash("Profile updated successfully.")

            return redirect(url_for("admin_profile"))

        # ------------------------------
        # Load Admin Details
        # ------------------------------

        cursor.execute("""
            SELECT
                id,
                name,
                username,
                email,
                profile_image
            FROM admins
            WHERE id=%s
        """,
        (
            1,
        ))

        admin = cursor.fetchone()

        return render_template(
            "admin_profile.html",
            admin=admin
        )

    except Exception as e:

        flash(f"An error occurred: {e}")

        return redirect(url_for("admin_dashboard"))

    finally:

        cursor.close()
        conn.close()


# ------------------------------
# ADD BOOK
# ------------------------------
@app.route("/admin/add-book", methods=["GET", "POST"])
def add_book():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    if request.method == "POST":

        name = request.form.get("name").strip()
        author = request.form.get("author").strip()
        category = request.form.get("category").strip()

        try:
            quantity = int(request.form.get("qty"))
        except:
            flash("Quantity must be a number.")
            return redirect(url_for("add_book"))

        if quantity < 0:
            flash("Quantity cannot be negative.")
            return redirect(url_for("add_book"))

        conn = get_db_connection()

        if conn is None:
            flash("Database connection failed.")
            return redirect(url_for("add_book"))

        cursor = conn.cursor()

        try:

            cursor.execute("""
                INSERT INTO books
                (book_name, author, category, quantity)
                VALUES(%s,%s,%s,%s)
            """, (name, author, category, quantity))

            conn.commit()
            
            log_activity(f"📚 New book added : {name}")

            flash("Book added successfully.")

            return redirect(url_for("view_books"))

        except Error as e:

            flash(f"Error : {e}")

        finally:

            cursor.close()
            conn.close()

    return render_template("add_book.html")


# ------------------------------
# VIEW / UPDATE / DELETE BOOKS
# ------------------------------

@app.route("/admin/view-books", methods=["GET", "POST"])
def view_books():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("admin_dashboard"))

    cursor = conn.cursor(dictionary=True)

    try:

        # ---------------------------------
        # UPDATE / DELETE
        # ---------------------------------

        if request.method == "POST":

            action = request.form.get("action")
            book_id = request.form.get("id")

            # ---------- UPDATE ----------

            if action == "update":

                name = request.form.get("name").strip()
                author = request.form.get("author").strip()
                category = request.form.get("category").strip()

                try:
                    quantity = int(request.form.get("qty"))
                except:
                    flash("Invalid quantity.")
                    return redirect(url_for("view_books"))

                if quantity < 0:
                    flash("Quantity cannot be negative.")
                    return redirect(url_for("view_books"))

                cursor.execute("""
                    UPDATE books
                    SET
                        book_name=%s,
                        author=%s,
                        category=%s,
                        quantity=%s
                    WHERE id=%s
                """,
                (
                    name,
                    author,
                    category,
                    quantity,
                    book_id
                ))

                conn.commit()

                flash("Book updated successfully.")

            # ---------- DELETE ----------

            elif action == "delete":

                cursor.execute("""
                    SELECT COUNT(*) AS total
                    FROM issued_books
                    WHERE
                        book_id=%s
                        AND status='Issued'
                """, (book_id,))

                active = cursor.fetchone()

                if active["total"] > 0:

                    flash("Book cannot be deleted because it is currently issued.")

                else:

                    cursor.execute("""
                        DELETE
                        FROM books
                        WHERE id=%s
                    """, (book_id,))

                    conn.commit()

                    log_activity("🗑️ Book deleted.")

                    flash("Book deleted successfully.")

            return redirect(url_for("view_books"))

        # ---------------------------------
        # SEARCH + PAGINATION
        # ---------------------------------

        page = request.args.get("page", 1, type=int)

        search = request.args.get("search", "").strip()

        per_page = 10

        offset = (page - 1) * per_page

        # ---------- COUNT BOOKS ----------

        if search:

            cursor.execute("""
                SELECT COUNT(*) AS total
                FROM books
                WHERE
                    book_name LIKE %s
                    OR author LIKE %s
                    OR category LIKE %s
            """,
            (
                f"%{search}%",
                f"%{search}%",
                f"%{search}%"
            ))

        else:

            cursor.execute("""
                SELECT COUNT(*) AS total
                FROM books
            """)

        total_books = cursor.fetchone()["total"]

        total_pages = (total_books + per_page - 1) // per_page

        # ---------- FETCH BOOKS ----------

        if search:

            cursor.execute("""
                SELECT *
                FROM books
                WHERE
                    book_name LIKE %s
                    OR author LIKE %s
                    OR category LIKE %s
                ORDER BY id
                LIMIT %s OFFSET %s
            """,
            (
                f"%{search}%",
                f"%{search}%",
                f"%{search}%",
                per_page,
                offset
            ))

        else:

            cursor.execute("""
                SELECT *
                FROM books
                ORDER BY id
                LIMIT %s OFFSET %s
            """,
            (
                per_page,
                offset
            ))

        books = cursor.fetchall()

        return render_template(
            "view_books.html",
            books=books,
            page=page,
            per_page=per_page,
            total_books=total_books,
            total_pages=total_pages,
            search=search
        )

    except Error as e:

        flash(f"Error : {e}")

        return redirect(url_for("admin_dashboard"))

    finally:

        cursor.close()
        conn.close()

# ------------------------------
# EXPORT BOOKS (CSV)
# ------------------------------
@app.route("/admin/export/books/csv")
def export_books_csv():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if search:

        cursor.execute("""
            SELECT
                id,
                book_name,
                author,
                category,
                quantity
            FROM books
            WHERE
                book_name LIKE %s
                OR author LIKE %s
                OR category LIKE %s
            ORDER BY id
        """,
        (
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
            SELECT
                id,
                book_name,
                author,
                category,
                quantity
            FROM books
            ORDER BY id
        """)

    books = cursor.fetchall()

    cursor.close()
    conn.close()

    import csv
    from io import StringIO

    output = StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "ID",
        "Book",
        "Author",
        "Category",
        "Quantity"
    ])

    for book in books:

        writer.writerow([
            book["id"],
            book["book_name"],
            book["author"],
            book["category"],
            book["quantity"]
        ])

    output.seek(0)

    return Response(

        output.getvalue(),

        mimetype="text/csv",

        headers={
            "Content-Disposition":
            "attachment; filename=library_books.csv"
        }

    )


# ------------------------------
# EXPORT BOOKS (EXCEL)
# ------------------------------
@app.route("/admin/export/books/excel")
def export_books_excel():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if search:

        cursor.execute("""
            SELECT
                id,
                book_name,
                author,
                category,
                quantity
            FROM books
            WHERE
                book_name LIKE %s
                OR author LIKE %s
                OR category LIKE %s
            ORDER BY id
        """,
        (
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
            SELECT
                id,
                book_name,
                author,
                category,
                quantity
            FROM books
            ORDER BY id
        """)

    books = cursor.fetchall()

    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Library Books"

    headers = [
        "ID",
        "Book",
        "Author",
        "Category",
        "Quantity"
    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    row = 2

    for book in books:

        ws.cell(row=row, column=1).value = book["id"]
        ws.cell(row=row, column=2).value = book["book_name"]
        ws.cell(row=row, column=3).value = book["author"]
        ws.cell(row=row, column=4).value = book["category"]
        ws.cell(row=row, column=5).value = book["quantity"]

        row += 1

    for column in ws.columns:

        length = max(len(str(cell.value or "")) for cell in column)

        ws.column_dimensions[
            get_column_letter(column[0].column)
        ].width = length + 4

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return Response(

        output.getvalue(),

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={
            "Content-Disposition":
            "attachment; filename=library_books.xlsx"
        }

    )

# ------------------------------
# EXPORT BOOKS (PDF)
# ------------------------------
@app.route("/admin/export/books/pdf")
def export_books_pdf():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if search:

        cursor.execute("""
            SELECT
                id,
                book_name,
                author,
                category,
                quantity
            FROM books
            WHERE
                book_name LIKE %s
                OR author LIKE %s
                OR category LIKE %s
            ORDER BY id
        """,
        (
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
            SELECT
                id,
                book_name,
                author,
                category,
                quantity
            FROM books
            ORDER BY id
        """)

    books = cursor.fetchall()

    cursor.close()
    conn.close()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4)
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<b><font size=18>Library Management System</font></b>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "<b>Books Report</b>",
            styles["Heading2"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Generated By :</b> {session['name']}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Date :</b> {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Total Books :</b> {len(books)}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph("<br/>", styles["Normal"])
    )

    table_data = [[
        "ID",
        "Book",
        "Author",
        "Category",
        "Quantity"
    ]]

    for book in books:

        table_data.append([
            book["id"],
            book["book_name"],
            book["author"],
            book["category"],
            book["quantity"]
        ])

    table = Table(table_data)

    table.setStyle(TableStyle([

        ("BACKGROUND", (0,0), (-1,0), colors.darkblue),

        ("TEXTCOLOR", (0,0), (-1,0), colors.white),

        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

        ("ALIGN", (0,0), (-1,-1), "CENTER"),

        ("GRID", (0,0), (-1,-1), 1, colors.grey),

        ("BACKGROUND", (0,1), (-1,-1), colors.beige),

        ("BOTTOMPADDING", (0,0), (-1,0), 10),

    ]))

    elements.append(table)

    doc.build(elements)

    pdf = buffer.getvalue()

    buffer.close()

    return Response(

        pdf,

        mimetype="application/pdf",

        headers={
            "Content-Disposition":
            "attachment; filename=library_books.pdf"
        }

    )

        
# ------------------------------
# ADD STUDENT
# ------------------------------
@app.route("/admin/add-student", methods=["GET", "POST"])
def add_student():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    if request.method == "POST":

        name = request.form.get("name").strip()
        student_class = request.form.get("class").strip()
        phone = request.form.get("phone").strip()
        password = request.form.get("password")

        # ---------- Validation ----------

        if len(name) < 2:
            flash("Student name is too short.")
            return redirect(url_for("add_student"))

        if not phone.isdigit() or len(phone) != 10:
            flash("Enter a valid 10-digit phone number.")
            return redirect(url_for("add_student"))

        if len(password) < 4:
            flash("Password must contain at least 4 characters.")
            return redirect(url_for("add_student"))

        hashed_password = password

        conn = get_db_connection()

        if conn is None:
            flash("Database connection failed.")
            return redirect(url_for("add_student"))

        cursor = conn.cursor()

        try:

            cursor.execute("""
                INSERT INTO students
                (name, class, phone, password)
                VALUES(%s,%s,%s,%s)
            """,
            (
                name,
                student_class,
                phone,
                hashed_password
            ))

            conn.commit()
            
            log_activity(f"👨 Student registered : {name}")

            flash("Student added successfully.")

            return redirect(url_for("view_students"))

        except Error as e:

            flash(f"Error : {e}")

        finally:

            cursor.close()
            conn.close()

    return render_template("add_student.html")


# ------------------------------
# VIEW / UPDATE / DELETE STUDENTS
# ------------------------------
@app.route("/admin/view-students", methods=["GET", "POST"])
def view_students():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("admin_dashboard"))

    cursor = conn.cursor(dictionary=True)

    try:

        # ---------------- UPDATE / DELETE ----------------

        if request.method == "POST":

            action = request.form.get("action")
            student_id = request.form.get("id")

            if action == "update":

                name = request.form.get("name").strip()
                student_class = request.form.get("class").strip()
                phone = request.form.get("phone").strip()
                password = request.form.get("password").strip()

                cursor.execute("""
                    UPDATE students
                    SET
                        name=%s,
                        class=%s,
                        phone=%s,
                        password=%s
                    WHERE id=%s
                """,
                (
                    name,
                    student_class,
                    phone,
                    password,
                    student_id
                ))

                conn.commit()

                flash("Student updated successfully.")

            elif action == "delete":

                cursor.execute("""
                    SELECT COUNT(*) AS total
                    FROM issued_books
                    WHERE
                        student_id=%s
                        AND status='Issued'
                """, (student_id,))

                active = cursor.fetchone()

                if active["total"] > 0:

                    flash("Student cannot be deleted because books are still issued.")

                else:

                    cursor.execute("""
                        DELETE FROM students
                        WHERE id=%s
                    """, (student_id,))

                    conn.commit()

                    log_activity("🗑️ Student deleted.")

                    flash("Student deleted successfully.")

            return redirect(url_for("view_students"))

        # ---------------- SEARCH + PAGINATION ----------------

        page = request.args.get("page", 1, type=int)
        search = request.args.get("search", "").strip()

        per_page = 10

        offset = (page - 1) * per_page

        # ---------------- COUNT ----------------

        if search:

            cursor.execute("""
                SELECT COUNT(*) AS total
                FROM students
                WHERE
                    id LIKE %s
                    OR name LIKE %s
                    OR class LIKE %s
                    OR phone LIKE %s
            """,
            (
                f"%{search}%",
                f"%{search}%",
                f"%{search}%",
                f"%{search}%"
            ))

        else:

            cursor.execute("""
                SELECT COUNT(*) AS total
                FROM students
            """)

        total_students = cursor.fetchone()["total"]

        total_pages = (total_students + per_page - 1) // per_page

        # ---------------- FETCH DATA ----------------

        if search:

            cursor.execute("""
                SELECT *
                FROM students
                WHERE
                    id LIKE %s
                    OR name LIKE %s
                    OR class LIKE %s
                    OR phone LIKE %s
                ORDER BY id
                LIMIT %s OFFSET %s
            """,
            (
                f"%{search}%",
                f"%{search}%",
                f"%{search}%",
                f"%{search}%",
                per_page,
                offset
            ))

        else:

            cursor.execute("""
                SELECT *
                FROM students
                ORDER BY id
                LIMIT %s OFFSET %s
            """,
            (
                per_page,
                offset
            ))

        students = cursor.fetchall()

        return render_template(
            "view_students.html",
            students=students,
            page=page,
            per_page=per_page,
            total_students=total_students,
            total_pages=total_pages,
            search=search
        )

    except Error as e:

        flash(f"Error : {e}")

        return redirect(url_for("admin_dashboard"))

    finally:

        cursor.close()
        conn.close()

# ------------------------------
# EXPORT STUDENTS CSV
# ------------------------------
@app.route("/admin/export/students/csv")
def export_students_csv():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if search:

        cursor.execute("""
            SELECT id,name,class,phone,password
            FROM students
            WHERE
                id LIKE %s
                OR name LIKE %s
                OR class LIKE %s
                OR phone LIKE %s
            ORDER BY id
        """,
        (
            f"%{search}%",
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
            SELECT id,name,class,phone,password
            FROM students
            ORDER BY id
        """)

    students = cursor.fetchall()

    cursor.close()
    conn.close()

    output = io.StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "ID",
        "Name",
        "Class",
        "Phone",
        "Password"
    ])

    for s in students:

        writer.writerow([
            s["id"],
            s["name"],
            s["class"],
            s["phone"],
            s["password"]
        ])

    response = make_response(output.getvalue())

    response.headers["Content-Disposition"] = \
        "attachment; filename=students_report.csv"

    response.headers["Content-Type"] = "text/csv"

    return response

# ------------------------------
# EXPORT STUDENTS EXCEL
# ------------------------------
@app.route("/admin/export/students/excel")
def export_students_excel():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if search:

        cursor.execute("""
            SELECT id,name,class,phone,password
            FROM students
            WHERE
                id LIKE %s
                OR name LIKE %s
                OR class LIKE %s
                OR phone LIKE %s
            ORDER BY id
        """,
        (
            f"%{search}%",
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
            SELECT id,name,class,phone,password
            FROM students
            ORDER BY id
        """)

    students = cursor.fetchall()

    cursor.close()
    conn.close()

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Students"

    headers = [
        "ID",
        "Name",
        "Class",
        "Phone",
        "Password"
    ]

    sheet.append(headers)

    for cell in sheet[1]:

        cell.font = Font(bold=True)

    for s in students:

        sheet.append([
            s["id"],
            s["name"],
            s["class"],
            s["phone"],
            s["password"]
        ])

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        download_name="students_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ------------------------------
# EXPORT STUDENTS PDF
# ------------------------------
@app.route("/admin/export/students/pdf")
def export_students_pdf():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if search:

        cursor.execute("""
            SELECT id,name,class,phone,password
            FROM students
            WHERE
                id LIKE %s
                OR name LIKE %s
                OR class LIKE %s
                OR phone LIKE %s
            ORDER BY id
        """,
        (
            f"%{search}%",
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
            SELECT id,name,class,phone,password
            FROM students
            ORDER BY id
        """)

    students = cursor.fetchall()

    cursor.close()
    conn.close()

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph("<b>Students Report</b>", styles["Title"])
    )

    elements.append(Spacer(1, 12))

    data = [[
        "ID",
        "Name",
        "Class",
        "Phone",
        "Password"
    ]]

    for s in students:

        data.append([
            s["id"],
            s["name"],
            s["class"],
            s["phone"],
            s["password"]
        ])

    table = Table(data)

    table.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.darkblue),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("GRID",(0,0),(-1,-1),1,colors.black),
        ("BOTTOMPADDING",(0,0),(-1,0),10),
        ("BACKGROUND",(0,1),(-1,-1),colors.beige)

    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="students_report.pdf",
        mimetype="application/pdf"
    )

# ------------------------------
# ISSUE BOOK
# ------------------------------
@app.route("/admin/issue-book", methods=["GET", "POST"])
def issue_book():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("admin_dashboard"))

    cursor = conn.cursor(dictionary=True)

    try:

        if request.method == "POST":

            student_id = request.form.get("student_id")
            book_id = request.form.get("book_id")

            # ----------------------------------
            # Automatic Date Generation
            # ----------------------------------

            issue_date = datetime.now().date()

            # Book will be due after 15 days
            return_date = issue_date + timedelta(days=15)

            # ----------------------------------
            # Manual Dates (Old Code)
            # Uncomment if you want manual dates.
            # ----------------------------------

            # issue_date = request.form.get("issue_date")
            # return_date = request.form.get("return_date")
            

            cursor.execute(
                "SELECT quantity FROM books WHERE id=%s",
                (book_id,)
            )

            book = cursor.fetchone()

            if not book:
                flash("Book not found.")
                return redirect(url_for("issue_book"))

            if book["quantity"] <= 0:
                flash("Book is out of stock.")
                return redirect(url_for("issue_book"))

            cursor.execute("""
                INSERT INTO issued_books
                (student_id, book_id, issue_date, return_date)
                VALUES(%s,%s,%s,%s)
            """,
            (
                student_id,
                book_id,
                issue_date,
                return_date
            ))

            cursor.execute("""
                UPDATE books
                SET quantity = quantity - 1
                WHERE id=%s
            """, (book_id,))

            conn.commit()
            
            log_activity("📤 A book was issued.")

            flash("Book issued successfully.")

            return redirect(url_for("view_issued_books"))

        cursor.execute("SELECT id, name FROM students")
        students = cursor.fetchall()

        cursor.execute("""
            SELECT id, book_name
            FROM books
            WHERE quantity > 0
        """)

        books = cursor.fetchall()

        return render_template(
            "issue_book.html",
            students=students,
            books=books
        )

    except Error as e:

        flash(str(e))

        return redirect(url_for("admin_dashboard"))

    finally:

        cursor.close()
        conn.close()


# ------------------------------
# VIEW ISSUED BOOKS
# ------------------------------
@app.route("/admin/view-issued")
def view_issued_books():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("admin_dashboard"))

    cursor = conn.cursor(dictionary=True)

    try:

        page = request.args.get("page", 1, type=int)
        search = request.args.get("search", "").strip()

        per_page = 10

        offset = (page - 1) * per_page

        # ---------------- COUNT ----------------

        if search:

            cursor.execute("""
                SELECT COUNT(*) AS total
                FROM issued_books i
                JOIN students s ON i.student_id=s.id
                JOIN books b ON i.book_id=b.id
                WHERE
                    s.name LIKE %s
                    OR b.book_name LIKE %s
                    OR i.status LIKE %s
            """,
            (
                f"%{search}%",
                f"%{search}%",
                f"%{search}%"
            ))

        else:

            cursor.execute("""
                SELECT COUNT(*) AS total
                FROM issued_books
            """)

        total_records = cursor.fetchone()["total"]

        total_pages = (total_records + per_page - 1) // per_page

        # ---------------- FETCH ----------------

        if search:

            cursor.execute("""
                SELECT
                    i.id,
                    s.name AS student_name,
                    b.book_name,
                    i.issue_date,
                    i.return_date,
                    i.status,
                    i.fine_amount
                FROM issued_books i
                JOIN students s
                    ON i.student_id=s.id
                JOIN books b
                    ON i.book_id=b.id
                WHERE
                    s.name LIKE %s
                    OR b.book_name LIKE %s
                    OR i.status LIKE %s
                ORDER BY i.id DESC
                LIMIT %s OFFSET %s
            """,
            (
                f"%{search}%",
                f"%{search}%",
                f"%{search}%",
                per_page,
                offset
            ))

        else:

            cursor.execute("""
                SELECT
                    i.id,
                    s.name AS student_name,
                    b.book_name,
                    i.issue_date,
                    i.return_date,
                    i.status,
                    i.fine_amount
                FROM issued_books i
                JOIN students s
                    ON i.student_id=s.id
                JOIN books b
                    ON i.book_id=b.id
                ORDER BY i.id DESC
                LIMIT %s OFFSET %s
            """,
            (
                per_page,
                offset
            ))

        issued = cursor.fetchall()

        today = datetime.now().date()

        for record in issued:

            if record["status"] == "Issued":

                if today > record["return_date"]:

                    fine = (today - record["return_date"]).days * 5

                    cursor.execute("""
                        UPDATE issued_books
                        SET fine_amount=%s
                        WHERE id=%s
                    """,
                    (
                        fine,
                        record["id"]
                    ))

                    record["fine_amount"] = fine

        conn.commit()

        return render_template(
            "view_issued.html",
            issued=issued,
            page=page,
            per_page=per_page,
            total_records=total_records,
            total_pages=total_pages,
            search=search
        )

    finally:

        cursor.close()
        conn.close()


# ------------------------------
# EXPORT ISSUED BOOKS CSV
# ------------------------------
@app.route("/admin/export/issued/csv")
def export_issued_csv():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if search:

        cursor.execute("""
            SELECT
                i.id,
                s.name AS student_name,
                b.book_name,
                i.issue_date,
                i.return_date,
                i.status,
                i.fine_amount
            FROM issued_books i
            JOIN students s ON i.student_id=s.id
            JOIN books b ON i.book_id=b.id
            WHERE
                s.name LIKE %s
                OR b.book_name LIKE %s
                OR i.status LIKE %s
            ORDER BY i.id DESC
        """,
        (
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
            SELECT
                i.id,
                s.name AS student_name,
                b.book_name,
                i.issue_date,
                i.return_date,
                i.status,
                i.fine_amount
            FROM issued_books i
            JOIN students s ON i.student_id=s.id
            JOIN books b ON i.book_id=b.id
            ORDER BY i.id DESC
        """)

    issued = cursor.fetchall()

    cursor.close()
    conn.close()

    output = io.StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "Sr. No.",
        "Student",
        "Book",
        "Issue Date",
        "Return Date",
        "Status",
        "Fine"
    ])

    for i in issued:

        writer.writerow([
            i["id"],
            i["student_name"],
            i["book_name"],
            i["issue_date"],
            i["return_date"],
            i["status"],
            i["fine_amount"]
        ])

    response = make_response(output.getvalue())

    response.headers["Content-Disposition"] = \
        "attachment; filename=issued_books_report.csv"

    response.headers["Content-Type"] = "text/csv"

    return response

# ------------------------------
# EXPORT ISSUED BOOKS EXCEL
# ------------------------------
@app.route("/admin/export/issued/excel")
def export_issued_excel():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if search:

        cursor.execute("""
            SELECT
                i.id,
                s.name AS student_name,
                b.book_name,
                i.issue_date,
                i.return_date,
                i.status,
                i.fine_amount
            FROM issued_books i
            JOIN students s
                ON i.student_id = s.id
            JOIN books b
                ON i.book_id = b.id
            WHERE
                s.name LIKE %s
                OR b.book_name LIKE %s
                OR i.status LIKE %s
            ORDER BY i.id DESC
        """,
        (
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
            SELECT
                i.id,
                s.name AS student_name,
                b.book_name,
                i.issue_date,
                i.return_date,
                i.status,
                i.fine_amount
            FROM issued_books i
            JOIN students s
                ON i.student_id = s.id
            JOIN books b
                ON i.book_id = b.id
            ORDER BY i.id DESC
        """)

    issued = cursor.fetchall()

    cursor.close()
    conn.close()

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Issued Books"

    headers = [
        "Sr. No.",
        "Student",
        "Book",
        "Issue Date",
        "Return Date",
        "Status",
        "Fine (INR)"
    ]

    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for record in issued:

        sheet.append([
            record["id"],
            record["student_name"],
            record["book_name"],
            record["issue_date"],
            record["return_date"],
            record["status"],
            record["fine_amount"]
        ])

    # Auto-size columns

    for column in sheet.columns:

        max_length = 0

        column_letter = get_column_letter(column[0].column)

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(str(cell.value))

            except:

                pass

        sheet.column_dimensions[column_letter].width = max_length + 3

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="issued_books_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ------------------------------
# EXPORT ISSUED BOOKS PDF
# ------------------------------
@app.route("/admin/export/issued/pdf")
def export_issued_pdf():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        if search:

            cursor.execute("""
                SELECT
                    i.id,
                    s.name AS student_name,
                    b.book_name,
                    i.issue_date,
                    i.return_date,
                    i.status,
                    i.fine_amount
                FROM issued_books i
                JOIN students s
                    ON i.student_id = s.id
                JOIN books b
                    ON i.book_id = b.id
                WHERE
                    s.name LIKE %s
                    OR b.book_name LIKE %s
                    OR i.status LIKE %s
                ORDER BY i.id DESC
            """,
            (
                f"%{search}%",
                f"%{search}%",
                f"%{search}%"
            ))

        else:

            cursor.execute("""
                SELECT
                    i.id,
                    s.name AS student_name,
                    b.book_name,
                    i.issue_date,
                    i.return_date,
                    i.status,
                    i.fine_amount
                FROM issued_books i
                JOIN students s
                    ON i.student_id = s.id
                JOIN books b
                    ON i.book_id = b.id
                ORDER BY i.id DESC
            """)

        issued = cursor.fetchall()

    finally:

        cursor.close()
        conn.close()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4)
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph("<b>Issued Books Report</b>", styles["Title"])
    )

    elements.append(Spacer(1, 12))

    data = [[
        "Sr. No.",
        "Student",
        "Book",
        "Issue Date",
        "Return Date",
        "Status",
        "Fine (INR)"
    ]]

    for record in issued:

        data.append([
            record["id"],
            record["student_name"],
            record["book_name"],
            str(record["issue_date"]),
            str(record["return_date"]),
            record["status"],
            record["fine_amount"]
        ])

    table = Table(data)

    table.setStyle(TableStyle([

        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 8),

    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="issued_books_report.pdf",
        mimetype="application/pdf"
    )

# ------------------------------
# RETURN BOOK
# ------------------------------
@app.route("/admin/return-book", methods=["GET", "POST"])
def return_book():

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("admin_dashboard"))

    cursor = conn.cursor(dictionary=True)

    try:

        if request.method == "POST":

            issued_id = request.form.get("issued_id")

            cursor.execute("""
                SELECT *
                FROM issued_books
                WHERE id=%s
            """,
            (issued_id,)
            )

            issue = cursor.fetchone()

            if not issue:

                flash("Record not found.")
                return redirect(url_for("return_book"))

            today = datetime.now().date()

            fine = 0

            if today > issue["return_date"]:
                fine = (today - issue["return_date"]).days * 5

            cursor.execute("""
                UPDATE issued_books
                SET
                    status='Returned',
                    fine_amount=%s
                WHERE id=%s
            """,
            (
                fine,
                issued_id
            ))

            cursor.execute("""
                UPDATE books
                SET quantity=quantity+1
                WHERE id=%s
            """,
            (
                issue["book_id"],
            ))

            conn.commit()
            
            log_activity("📥 A book was returned.")

            flash("Book returned successfully.")

            return redirect(url_for("view_issued_books"))

        cursor.execute("""
        SELECT
            i.id,
            s.name,
            b.book_name
        FROM issued_books i
        JOIN students s ON i.student_id=s.id
        JOIN books b ON i.book_id=b.id
        WHERE i.status='Issued'
        """)

        active_issues = cursor.fetchall()

        return render_template(
            "return_book.html",
            active_issues=active_issues
        )

    finally:

        cursor.close()
        conn.close()


# ------------------------------
# STUDENT DASHBOARD
# ------------------------------
@app.route("/student/dashboard")
def student_dashboard():

    if session.get("role") != "student":
        return redirect(url_for("login_page"))

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("login_page"))

    cursor = conn.cursor(dictionary=True)

    try:

        # ---------------------------------
        # Pagination + Search
        # ---------------------------------

        page = request.args.get("page", 1, type=int)

        search = request.args.get("search", "").strip()

        per_page = 10

        offset = (page - 1) * per_page

        # ---------------------------------
        # Dashboard Statistics
        # ---------------------------------

        cursor.execute("""
            SELECT

                SUM(CASE WHEN status='Issued' THEN 1 ELSE 0 END) AS borrowed_books,

                SUM(CASE WHEN status='Returned' THEN 1 ELSE 0 END) AS returned_books,

                SUM(
                    CASE
                        WHEN status='Issued'
                        AND fine_status='Pending'
                        THEN 1
                        ELSE 0
                    END
                ) AS overdue_books,

                SUM(
                    CASE
                        WHEN fine_status='Pending'
                        THEN fine_amount
                        ELSE 0
                    END
                ) AS outstanding_fine

            FROM issued_books

            WHERE student_id=%s
        """,
        (
            session["user_id"],
        ))

        stats = cursor.fetchone()

        # ---------------------------------
        # Total Books (for pagination)
        # ---------------------------------

        if search:

            cursor.execute("""
                SELECT COUNT(*) AS total
                FROM issued_books i
                JOIN books b
                ON i.book_id=b.id
                WHERE
                    i.student_id=%s
                    AND
                    (
                        b.book_name LIKE %s
                        OR b.author LIKE %s
                    )
            """,
            (
                session["user_id"],
                f"%{search}%",
                f"%{search}%"
            ))

        else:

            cursor.execute("""
                SELECT COUNT(*) AS total
                FROM issued_books
                WHERE student_id=%s
            """,
            (
                session["user_id"],
            ))

        total_books = cursor.fetchone()["total"]

        total_pages = (total_books + per_page - 1) // per_page

        # ---------------------------------
        # Fetch Books
        # ---------------------------------

        if search:

            cursor.execute("""
                SELECT
                    i.id,
                    b.cover_image,
                    b.book_name,
                    b.author,
                    i.issue_date,
                    i.return_date,
                    i.status,
                    i.fine_amount,
                    i.fine_status
                FROM issued_books i
                JOIN books b
                ON i.book_id=b.id
                WHERE
                    i.student_id=%s
                    AND
                    (
                        b.book_name LIKE %s
                        OR b.author LIKE %s
                    )
                ORDER BY i.id DESC
                LIMIT %s OFFSET %s
            """,
            (
                session["user_id"],
                f"%{search}%",
                f"%{search}%",
                per_page,
                offset
            ))

        else:

            cursor.execute("""
                SELECT
                    i.id,
                    b.cover_image,
                    b.book_name,
                    b.author,
                    i.issue_date,
                    i.return_date,
                    i.status,
                    i.fine_amount,
                    i.fine_status
                FROM issued_books i
                JOIN books b
                ON i.book_id=b.id
                WHERE i.student_id=%s
                ORDER BY i.id DESC
                LIMIT %s OFFSET %s
            """,
            (
                session["user_id"],
                per_page,
                offset
            ))

        books = cursor.fetchall()

        # ---------------------------------
        # Update Overdue Fines
        # ---------------------------------

        cursor = conn.cursor(dictionary=True)
        
        today = datetime.now().date()
        
        cursor.execute("""
        SELECT
            id,
            return_date,
            status
        FROM issued_books
        WHERE
            student_id=%s
            AND status='Issued'
        """,
        (
            session["user_id"],
        ))
        
        issued_books = cursor.fetchall()
        
        for book in issued_books:
            if today > book["return_date"]:
                fine = (today - book["return_date"]).days * 5
                
                cursor.execute("""
                UPDATE issued_books
                SET
                    fine_amount=%s,
                    fine_status='Pending'
                WHERE id=%s
                """,
                (
                    fine,
                    book["id"]
                ))
                
        conn.commit()        

        return render_template(
            "student_dashboard.html",
            name=session["name"],
            books=books,
            stats=stats,
            page=page,
            per_page=per_page,
            total_books=total_books,
            total_pages=total_pages,
            search=search
        )

    finally:

        cursor.close()
        conn.close()


# ------------------------------
# STUDENT PROFILE
# ------------------------------
@app.route("/student/profile", methods=["GET", "POST"])
def student_profile():

    if session.get("role") != "student":
        return redirect(url_for("login_page"))

    student_id = session.get("user_id")

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("student_dashboard"))

    cursor = conn.cursor(dictionary=True)

    try:

        if request.method == "POST":

            name = request.form.get("name").strip()
            email = request.form.get("email").strip()

            image = request.files.get("profile_image")

            current_password = request.form.get("current_password", "").strip()
            new_password = request.form.get("new_password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()

            if not name or not email:

                flash("All fields are required.")
                return redirect(url_for("student_profile"))

            # ------------------------------
            # Upload Profile Picture
            # ------------------------------

            filename = None

            if image and image.filename != "":

                extension = image.filename.rsplit(".", 1)[1].lower()

                filename = f"{student_id}.{extension}"

                image.save(
                    os.path.join(
                        app.static_folder,
                        "uploads",
                        "profiles",
                        filename
                    )
                )

            # ------------------------------
            # Change Password
            # ------------------------------

            password = None

            if current_password or new_password or confirm_password:

                if not current_password or not new_password or not confirm_password:

                    flash("Please fill all password fields.")
                    return redirect(url_for("student_profile"))

                cursor.execute("""
                    SELECT password
                    FROM students
                    WHERE id=%s
                """,
                (
                    student_id,
                ))

                student_data = cursor.fetchone()

                if student_data["password"] != current_password:

                    flash("Current password is incorrect.")
                    return redirect(url_for("student_profile"))

                if new_password != confirm_password:

                    flash("New password and confirm password do not match.")
                    return redirect(url_for("student_profile"))

                password = new_password

            # ------------------------------
            # Update Profile
            # ------------------------------

            if filename and password:

                cursor.execute("""
                    UPDATE students
                    SET
                        name=%s,
                        email=%s,
                        password=%s,
                        profile_image=%s
                    WHERE id=%s
                """,
                (
                    name,
                    email,
                    password,
                    filename,
                    student_id
                ))

            elif filename:

                cursor.execute("""
                    UPDATE students
                    SET
                        name=%s,
                        email=%s,
                        profile_image=%s
                    WHERE id=%s
                """,
                (
                    name,
                    email,
                    filename,
                    student_id
                ))

            elif password:

                cursor.execute("""
                    UPDATE students
                    SET
                        name=%s,
                        email=%s,
                        password=%s
                    WHERE id=%s
                """,
                (
                    name,
                    email,
                    password,
                    student_id
                ))

            else:

                cursor.execute("""
                    UPDATE students
                    SET
                        name=%s,
                        email=%s
                    WHERE id=%s
                """,
                (
                    name,
                    email,
                    student_id
                ))

            conn.commit()

            session["name"] = name

            if password:

                flash("Profile and password updated successfully.")

            else:

                flash("Profile updated successfully.")

            return redirect(url_for("student_profile"))

        # ------------------------------
        # Load Student Details
        # ------------------------------

        cursor.execute("""
            SELECT
                id,
                name,
                class,
                phone,
                email,
                profile_image
            FROM students
            WHERE id=%s
        """,
        (
            student_id,
        ))

        student = cursor.fetchone()

        return render_template(
            "student_profile.html",
            student=student
        )

    except Exception as e:

        flash(f"An error occurred: {e}")
        return redirect(url_for("student_dashboard"))

    finally:

        cursor.close()
        conn.close()

# ------------------------------
# UPLOAD BOOK COVER
# ------------------------------
@app.route("/admin/upload-cover/<int:book_id>", methods=["GET", "POST"])
def upload_book_cover(book_id):

    if session.get("role") != "admin":
        return redirect(url_for("login_page"))

    conn = get_db_connection()

    if conn is None:
        flash("Database connection failed.")
        return redirect(url_for("view_books"))

    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute(
            "SELECT * FROM books WHERE id=%s",
            (book_id,)
        )

        book = cursor.fetchone()

        if not book:

            flash("Book not found.")
            return redirect(url_for("view_books"))

        if request.method == "POST":

            file = request.files.get("cover")

            if not file or file.filename == "":

                flash("Please choose an image.")
                return redirect(request.url)

            if not allowed_file(file.filename):

                flash("Only JPG, JPEG, PNG and WEBP images are allowed.")
                return redirect(request.url)

            extension = file.filename.rsplit(".", 1)[1].lower()

            filename = f"{book_id}.{extension}"

            # Delete old cover if extension changed

            for ext in ALLOWED_EXTENSIONS:

                old_file = os.path.join(
                    app.config["UPLOAD_FOLDER"],
                    f"{book_id}.{ext}"
                )

                if os.path.exists(old_file):

                    os.remove(old_file)

            filepath = os.path.join(
                app.config["UPLOAD_FOLDER"],
                filename
            )

            file.save(filepath)

            cursor.execute(
                """
                UPDATE books
                SET cover_image=%s
                WHERE id=%s
                """,
                (
                    filename,
                    book_id
                )
            )

            conn.commit()

            flash("Book cover uploaded successfully.")

            return redirect(url_for("view_books"))

        return render_template(
            "upload_cover.html",
            book=book
        )

    finally:

        cursor.close()
        conn.close()

# ------------------------------
# EXPORT STUDENT BOOKS (CSV)
# ------------------------------
@app.route("/student/export/csv")
def export_student_csv():

    if session.get("role") != "student":
        return redirect(url_for("login_page"))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            b.book_name,
            b.author,
            i.issue_date,
            i.return_date,
            i.status,
            i.fine_amount
        FROM issued_books i
        JOIN books b
            ON i.book_id = b.id
        WHERE i.student_id=%s
        ORDER BY i.id DESC
    """,
    (
        session["user_id"],
    ))

    books = cursor.fetchall()

    cursor.close()
    conn.close()

    import csv
    from io import StringIO
    from flask import Response

    output = StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "Book",
        "Author",
        "Issue Date",
        "Return Date",
        "Status",
        "Fine"
    ])

    for book in books:

        writer.writerow([
            book["book_name"],
            book["author"],
            book["issue_date"],
            book["return_date"],
            book["status"],
            book["fine_amount"]
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition":
            "attachment; filename=my_books.csv"
        }
    )

# ------------------------------
# EXPORT STUDENT BOOKS (EXCEL)
# ------------------------------
@app.route("/student/export/excel")
def export_student_excel():

    if session.get("role") != "student":
        return redirect(url_for("login_page"))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            b.book_name,
            b.author,
            i.issue_date,
            i.return_date,
            i.status,
            i.fine_amount
        FROM issued_books i
        JOIN books b
            ON i.book_id = b.id
        WHERE i.student_id=%s
        ORDER BY i.id DESC
    """,
    (
        session["user_id"],
    ))

    books = cursor.fetchall()

    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "My Books"

    headers = [
        "Book",
        "Author",
        "Issue Date",
        "Return Date",
        "Status",
        "Fine"
    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=1, column=col)

        cell.value = header
        cell.font = Font(bold=True)

    row = 2

    for book in books:

        ws.cell(row=row, column=1).value = book["book_name"]
        ws.cell(row=row, column=2).value = book["author"]
        ws.cell(row=row, column=3).value = str(book["issue_date"])
        ws.cell(row=row, column=4).value = str(book["return_date"])
        ws.cell(row=row, column=5).value = book["status"]
        ws.cell(row=row, column=6).value = book["fine_amount"]

        row += 1

    for column in ws.columns:

        length = max(len(str(cell.value or "")) for cell in column)

        ws.column_dimensions[
            get_column_letter(column[0].column)
        ].width = length + 4

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return Response(

        output.getvalue(),

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={
            "Content-Disposition":
            "attachment; filename=my_books.xlsx"
        }

    )

# ------------------------------
# EXPORT STUDENT BOOKS (PDF)
# ------------------------------
@app.route("/student/export/pdf")
def export_student_pdf():

    if session.get("role") != "student":
        return redirect(url_for("login_page"))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            b.book_name,
            b.author,
            i.issue_date,
            i.return_date,
            i.status,
            i.fine_amount
        FROM issued_books i
        JOIN books b
            ON i.book_id=b.id
        WHERE i.student_id=%s
        ORDER BY i.id DESC
    """,
    (
        session["user_id"],
    ))

    books = cursor.fetchall()

    cursor.close()
    conn.close()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4)
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<b><font size=18>Library Management System</font></b>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Student :</b> {session['name']}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Generated :</b> {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph("<br/>", styles["Normal"])
    )

    total_fine = sum(book["fine_amount"] for book in books)

    elements.append(
        Paragraph(
            f"<b>Total Books :</b> {len(books)}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Outstanding Fine :</b> INR {total_fine}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph("<br/>", styles["Normal"])
    )

    table_data = [[
        "Book",
        "Author",
        "Issue Date",
        "Return Date",
        "Status",
        "Fine (INR)"
    ]]

    for book in books:

        table_data.append([
            book["book_name"],
            book["author"],
            str(book["issue_date"]),
            str(book["return_date"]),
            book["status"],
            f"{book['fine_amount']}"
        ])

    table = Table(table_data)

    table.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.darkblue),

        ("TEXTCOLOR",(0,0),(-1,0),colors.white),

        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ("ALIGN",(0,0),(-1,-1),"CENTER"),

        ("GRID",(0,0),(-1,-1),1,colors.grey),

        ("BACKGROUND",(0,1),(-1,-1),colors.beige),

        ("BOTTOMPADDING",(0,0),(-1,0),10),

    ]))

    elements.append(table)

    doc.build(elements)

    pdf = buffer.getvalue()

    buffer.close()

    return Response(

        pdf,

        mimetype="application/pdf",

        headers={

            "Content-Disposition":
            "attachment; filename=my_books.pdf"

        }

    )

# ------------------------------
# MAIN
# ------------------------------
if __name__ == "__main__":
    app.run(debug=False)        